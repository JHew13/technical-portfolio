# Sentence Similarity Analysis
# Compares workflow text to identify exact and partial similarities and potential duplicates.
#
# Portfolio note: This public sample is based on code I originally wrote for an
# internal business tool. Proprietary names, endpoints, identifiers, sample data,
# and environment-specific values have been replaced with generic equivalents.
from difflib import SequenceMatcher
from re import L
import pandas as pd
import itertools 

from itertools import permutations,product,zip_longest
import sqlalchemy as sa
from sqlalchemy import inspect 
from openpyxl import load_workbook

mapping ={}
wb = load_workbook("INPUT_WORKBOOK.xlsx")
ws = wb["Questions + HP"]
for entry, data_boundary in ws.tables.items():
    data = ws[data_boundary]
    content = [[cell.value for cell in ent]
    for ent in data ]

header = content[0]
rest = content[1:]
df = pd.DataFrame(rest,columns = header)
    
Table = mapping[entry]=df

healthPlans = []
#print (Table)
engine = sa.create_engine('mssql+pyodbc://SERVER_NAME/DATABASE_NAME?driver=SQL Server Native Client 11.0')

# In future would like to have this auto populate list with HP
# #df = pd.read_csv("HP_questions_python.csv", engine ='python')
 
#correctedSentence = df.values.tolist()
#prodcutDf = pd.DataFrame(correctedSentence(list(IT.product(*correctedSentence)),columns=['index','hp','question']))
filteredQuestion = list(filter(None, Table["Questions"])) # Isolates the questions from the other columns for easy access
questions = ['HCSC, ',
'HCSC, Is there a reason for the out of network site?',
'HCSC, Is there a reason for the out of network lab?',
'HCSC, Do you have a referral from a PCP?',
'HCSC, Has the specimen been collected?',
'HCSC, Have you obtained an OON waiver?',
'HCSC, The health plan provider you have selected may or may not be certified by your health plan to perform the requested service. I can provide you with a list of approved facilities for your requested service. Would you like a list of facilities located near the patient’s home or work?',
'HCSC, The provider you have selected is out of network.  Would you like to find an in-network provider?',
'Horizon / HORIZON MEDICAID, ',
'Horizon / HORIZON MEDICAID, Is there a reason for the out of network site?',
'Horizon / HORIZON MEDICAID, Is there a reason for the out of network lab?',
'Horizon / HORIZON MEDICAID, Do you have a referral from a PCP?',
'Horizon / HORIZON MEDICAID, Has the specimen been collected?',
'Horizon / HORIZON MEDICAID, Have you obtained an OON waiver?',
'Horizon / HORIZON MEDICAID, The health plan provider you have selected may or may not be certified by your health plan to perform the requested service. I can provide you with a list of approved facilities for your requested service. Would you like a list of facilities located near the patient’s home or work?',
'Horizon / HORIZON MEDICAID, The provider you have selected is out of network.  Would you like to find an in-network provider?',
'HAMP/The Health Plan, ',
'HAMP/The Health Plan, Is there a reason for the out of network site?',
'HAMP/The Health Plan, Is there a reason for the out of network lab?',
'HAMP/The Health Plan, Do you have a referral from a PCP?',
'HAMP/The Health Plan, Has the specimen been collected?',
'HAMP/The Health Plan, Have you obtained an OON waiver?',
'HAMP/The Health Plan, The health plan provider you have selected may or may not be certified by your health plan to perform the requested service. I can provide you with a list of approved facilities for your requested service. Would you like a list of facilities located near the patient’s home or work?',
'HAMP/The Health Plan, The provider you have selected is out of network.  Would you like to find an in-network provider?',
'Priority Health, ',
'Priority Health, Is there a reason for the out of network site?',
'Priority Health, Is there a reason for the out of network lab?',
'Priority Health, Do you have a referral from a PCP?',
'Priority Health, Has the specimen been collected?',
'Priority Health, Have you obtained an OON waiver?',
'Priority Health, The health plan provider you have selected may or may not be certified by your health plan to perform the requested service. I can provide you with a list of approved facilities for your requested service. Would you like a list of facilities located near the patient’s home or work?',
'Priority Health, The provider you have selected is out of network.  Would you like to find an in-network provider?',
'Health Plan D, ',
'Health Plan D, Is there a reason for the out of network lab?',
'Health Plan C/Health Plan A Plus Health Plan C, ',
'Health Plan C/Health Plan A Plus Health Plan C, Our records indicate that this is a provider or site is not contracted for this member\'sline of business. The site you have selected may result in a higher cost to the member. Would you like to proceed with this provider?',
'Health Plan C/Health Plan A Plus Health Plan C, Our records indicate that this is a provider or site is not contracted for this member\'sline of business. Would you like to select an in network site?',
'BCBSMN, The provider you have selected is out of network.  Would you like to find an in-network provider?',
'BCBSMN, ',
'BCBSMN, Is there a reason for the out of network lab?',
'BCBSMN, Is there a reason for the out of network site?',
'Scott and White, ',
'Scott and White, Our records indicate that this is a provider or site that is not contracted for this member\'s line of business.  Would you like to select an in network site?',
'Health Plan E, ',
'Health Plan E, This member does not have OON benefits, may I assist you in locating an INN lab?',
'Health Plan E, Is there a reason for the out of network lab?',
'Health Plan E, The site you have selected is out of network.  May I assist you in locating an INN lab?',
'Health Plan E, Our records indicate that this is a provider or site that is not contracted for this member\'s line of business.  Would you like to select an in network site?',
'Health Plan E, Our records indicate that this is a provider or site that is not participating for this member\'s line of business.  Would you like to select an in network site?',
'Health Plan E, What is your reason for going to the out of network provider?',
'Health Plan A, ',
'Health Plan A, DO NOT READ TO CALLER: Did the caller mention NAP coverage?',
'Health Plan F, ',
'Health Plan F, The DME supplier chosen is out of network.  The member is enrolled in a plan that allows OON supplier selection. By choosing an OON supplier, the member is at a higher financial risk. Would you like to proceed with the OON supplier or would you like to change to an in network supplier?',
'Health Plan F, The DME supplier chosen is out of network.  Would you like assistance in finding an in network supplier?',
'Highmark, ',
'Highmark, Our records indicate that this is a provider or site not contracted for this member\'sline of business.  The site you have selected may result in a higher cost to the member.  Would you like to select another site?',
'Highmark, This is a manually entered site, would you like assistance in finding an in-network privileged site?',
'Highmark, This is a manually entered site, would you like assistance in finding an in-network site?',
'Highmark, What is the reason for selecting an OON site?',
'Highmark, Based on our records we are unable to confirm network participation at this time. Based on the site you have selected, the member may experience higher cost sharing, would you like to select another site?',
'Moda/Summit, ',
'Moda/Summit, Our records indicate that this is a provider or site that does not participate with <InsCarrier> or is not contracted for this member\'sline of business.  Would you like to choose an in network provider or site?',
'Moda/Summit, Do you have a referral from the PCP in hand?',
'Moda/Summit, Do you have an in-network exception?',
'United, ',
'United, Our records indicate that this is a provider or site that does not participate with <InsCarrier> or is not contracted for this member\'sline of business.  Would you like to choose an in network provider or site?',
'United, Do you have a referral from the PCP in hand?',
'United, Do you have an in-network exception?',
'BCBSAZ, Our records indicate that this provider or site does not participate with BCBS AZ. This member\'sbenefit requires service at participating facility. Would you like to select an in network site?',
'BCBSAZ, Is there a reason for the out of network lab?',
'BCBSAZ, Is there a reason for the out of network site?',
'BCBSAZ, ',
'BCBSAZ, Our records indicate that this provider or site does not participate with BCBSAZ. This member\'sbenefit requires service at participating facility. Would you like to select an in network site?',
'Spreemo, ',
'Spreemo, Did you select this site due to one of the following medical exceptions?',
'ConnectiCare/BCBSND, ',
'ConnectiCare/BCBSND, This provider is not participating with the plan and may result in out of pocket costs to the member.  Would you like to do another search?',
'ConnectiCare/BCBSND, Our records indicate that this is a site that is not contracted for this member\'sline of business. The site you have selected may result in a higher cost to the member. Would you like to search again for an in network site?',
'ConnectiCare/BCBSND, Is there a reason for the selection of an OON site?',
'Providence, Our records indicate that this is a provider or site is not contracted for this member\'sline of business. Would you like to select an in network site?',
'Providence, Do any of the following conditions apply?',
'Providence, ',
'Providence, This member\'sbenefit requires service at a participating facility.  Our records indicate that this is a provider that does not participate with Providence Health Plan.  The site you have selected may result in a higher cost to the member.',
'Blue Care Network, Our records indicate this provider or site is not contracted for this member\'s line of business.  Would you like to search for an in-network provider for the member?',
'Blue Care Network, Please click Submit.',
'Blue Care Network, ',
'Health Plan B Networking, ',
'Health Plan B Networking, Our records indicate that this provider or site is not contracted for this member\'sline of business. The site you have selected may result in a higher cost to the member. Would you like to continue with this site or search again?',
'Health Plan B Networking, This Out of Network request does not require Prior Authorization through Example Service Company. Please contact Health Plan B on the back of the member\'s card to determine whether Prior Auth is required through Health Plan B.',
'Health Plan B Networking, Our records indicate that this provider or site is not found for this member\'s network. Please contact Example Service Company healthcare at (555) 010-0000 option 2 to confirm provider participation status.',
'Health Plan B Networking, Please Click Submit',
'Health Plan B Networking, The provider you have selected is not certified or accredited by the member\'shealth plan to perform the requested service. Please attempt to search by provider group, name, or TIN. Do not read to caller: To proceed with this site, select No and this case will route to E status. Select Yes to search again. If NO, READ: Once case creation is completed, a member of our Eligibility department will review the case and contact your office.',
'Sunshine Health, This member\'sbenefit requires service at a participating facility. Our records indicate that this is a provider or site that does not participate with the health plan. The site you have selected may result in a higher cost to the member. Would you like to proceed?',
'Sunshine Health, Does the patient have a need for continuity of care?',
'Meridian, Our records indicate that this provider or site is not contracted for this member\'s line of business. Would you like to select an in network site?',
'Meridian, Our records indicate that this provider or site is not contracted for this member\'sline of business. There are many sites who perform the same or similar tests that are contracted such as LabCorp or Quest Diagnostics.  Would you like to search again?',
'Meridian, Is there a reason for the selection of an OON lab?',
'Meridian, Is there a reason for the out of network site?',
'Meridian, ',
'HealthFirst Florida/Advent, Please click Submit.',
'HealthFirst Florida/Advent, ',
'HealthFirst Florida/Advent, Our records indicate this provider or site is not contracted for this member\'s line of business. The site you have selected may result in a higher cost to the member. Would you like to search for an in-network provider for the member?',
'HealthFirst Florida/Advent, Please select a reason for the use of OON lab',
'HealthFirst Florida/Advent, Is there a reason for the out of network site? Please select one of the following options:',
'HealthFirst Florida/Advent, Please click Submit. ',
'HealthFirst Florida/Advent, Our records indicate that this is a provider or site is not contracted for this member\'sline of business. The site you have selected may result in a higher cost to the member. Would you like to proceed with this provider?',
'HealthFirst Florida/Advent, Our records indicate that this is a provider or site is not contracted for this member\'sline of business. Would you like to select an in network site?',
'SummaCare, The site you have selected is a non-network facility and may result in higher costs to the member. Would you like to review alternative in-network facilities? Would you like to search again?',
'SummaCare, The site you have selected is not contracted for this member\'s line of business. Do you wish to continue with this rendering site?',
'SummaCare, Is this an Urgent Request?',
'SummaCare, Is there a reason for the out of network site? Please select one of the following options:',
'SummaCare, ',
'SummaCare, Is this an Urgent Request or Medicare MedOnc?',
'Clover Health, This site you have selected is not in network for this member\'s coverage.  Would you like to proceed with this request using this site?',
'Clover Health, Select out of network reason:',
'First Care, The site you have requested may result in a higher cost to the member.  Would you like to proceed?',
'First Care, This member\'sbenefit requires service at a participating facility. Our records indicate that this is a provider or site that does not participate with Health Plan. The site you have selected may result in a higher cost to the member. Would you like to proceed?',
'Health Partners Plan, This member\'s benefit requires service at a participating facility. Our records indicate that this is a provider or site that does not participate with Health Partners Plans. The site you have selected may result in a higher cost to the member. Would you like to proceed?',
'Health Partners Plan, ',
'Health Partners Plan, Select out of network reason:',
'Health Partners Plan, Is there a reason for the out of network site? Please select one of the following options:',
'Humana KY, This member\'s benefit requires service at a participating facility. Our records indicate that this is a provider or site that does not participate with the Health Plan. The site you have selected may result in a higher cost to the member. Would you like to proceed?',
'Humana KY, Is there a reason for the out of network site?  Please select one of the following options:',
'IBC, The site you have you have selected may result in a higher cost to the member. Would you like to proceed?',
'IBC, This member\'s benefit requires service at a participating facility. Our records indicate that this is a provider or site that does not participate with Health Plan. The site you have selected may result in a higher cost to the member. Would you like to proceed?',
'IBC, Select out of network reason:',
'IBC, ',
'Amerihealth/Independence Admin, The site you have you have selected may result in a higher cost to the member. Would you like to proceed?',
'Amerihealth/Independence Admin, This member\'s benefit requires service at a participating facility. Our records indicate that this is a provider or site that does not participate with Health Plan. The site you have selected may result in a higher cost to the member. Would you like to proceed?',
'Amerihealth/Independence Admin, Select out of network reason:',
'Amerihealth/Independence Admin, ',
'MMOH, Do any of the following conditions apply?',
'MMOH, The site you have selected may result in a higher cost to the member. Would you like to proceed?',
'MMOH, The site selected is not covered under the member\'s benefit. Prior authorization approval will not be granted unless BOTH medical necessity and benefit criteria are met. Proceeding with an out of network site may delay the prior authorization decision. Would you like to proceed? Please be advised that proceeding does not guarantee payment for services.',
'First Carolina Care, This member\'s benefit requires service at a participating facility. Our records indicate that this is a provider or site that does not participate with Health Plan. The site you have selected may result in a higher cost to the member. Would you like to proceed?',
'First Carolina Care, The site you have selected may result in a higher cost to the member. Would you like to proceed?',
'First Carolina Care, You have selected a facility which will require health plan review of participation status and member benefits. Final determination will be made by the health plan. Would you like to proceed?',
'First Carolina Care, Is there a reason for the out of network site? Please select one of the following options:',
'HMSA, This member\'s benefit requires service at a participating facility. Our records indicate that this is a provider or site that does not participate with Health Plan. The site you have selected may result in a higher cost to the member. Would you like to proceed?',
'HMSA, The site you have selected may result in a higher cost to the member. Would you like to proceed?',
'HMSA, Is there a reason for the out of network site? Please select one of the following options:',
'HMSA, ',
'Sanford, The site you have selected may result in a higher cost to the member. Would you like to proceed?',
'Sanford, This member\'s benefit requires service at a participating facility. Our records indicate that this is a provider or site that does not participate with Health Plan. The site you have selected may result in a higher cost to the member. Would you like to proceed?',
'Sanford, Is there a reason for the out of network site?  Please select one of the following options:',
'Johns Hopkins, This member\'s benefit requires service at a participating facility. Our records indicate that this is a provider or site that does not participate with Health Plan. The site you have selected may result in a higher cost to the member. Would you like to proceed?',
'Johns Hopkins, The site you have selected may result in a higher cost to the member. Would you like to proceed?',
'Johns Hopkins, Is there a reason for the OON Site? Please choose one of the following options:',
'Johns Hopkins, Is there a reason for the out of network site?  Please select one of the following options:',
'Arkansas BCBS, The DME Supplier you have chosen is not participating for this member\'s network. To receive the best benefit, the member should see care with an in-network provider. If you go to a provider outside of the network, the out-of-network benefits will apply with a high Out of Pocket cost and will accumulate toward the out-of-network deductible and out-of-pocket maximum. Would you like to continue with this site?',
'Arkansas BCBS, Our records indicate that this DME Supplier is not contracted for the member\'sline of business and this member does not have out of network coverage. If you continue with this site, additional review will be required. Would you like to continue with this site? ',
'HealthFirst Lab, This member\'s benefit requires service at a participating facility. Our records indicate that this is a provider or site that does not participate with Health Plan. The site you have selected may result in a higher cost to the member. Would you like to proceed?',
'HealthFirst Lab, Select out of network reason:',
'AultCare, This member\'s benefit requires service at a participating facility. Our records indicate that this is a provider or site that does not participate with Health Plan. The site you have selected may result in a higher cost to the member. Would you like to proceed?',
'AultCare, Please select a reason for the use of OON lab',
'AultCare, Is there a reason for the out of network site?  Please select one of the following options:',
'AultCare, ',
'Wellmed, ',
'Avmed, The site you have selected may result in a higher cost to the member. Would you like to proceed?',
'Avmed, This member\'s benefit requires service at a participating facility. Our records indicate that this is a provider or site that does not participate with Health Plan. The site you have selected may result in a higher cost to the member. Would you like to proceed?',
'Avmed, Is there a reason for the out of network site?  Please select one of the following options:',

]
filteredHP = list(filter(None, Table["HP"]))# Isolates the HP from the other columns for easy access
com = [',']
filtered = filteredHP + filteredQuestion

for f in filteredQuestion:
    filtered = filteredHP.append(filteredQuestion)
#print (filtered)
a = []
b =  []
s = []
e = []
originalList = []
ListCompared = []
hp = []
hpCompared =[]
question=[]
QuestionCompared = []
d = []
data = []
l=[]
percent=[]
test2=[]
#res = [v for v in zip_longest(*[k.split(', ') for k in questions])] this does something..


for a,b in permutations(questions,2):
    a = a.split(',')
    hp = a[0]
    question = a[1]
    my_list = list(hp)
    my_list.append(question)
    b = b.split(',')
    hpCompared = b[0]
    QuestionCompared = b[1]
    
    l = SequenceMatcher(None,question,QuestionCompared).ratio() 
    
    
    #print (s)
        
    #originalList.append(questions)
    
        
 

            
        

        #e = s * 100
       
        #c = a,b,e

    data.append({"HP":hp,
                "Questions":question,
                "ComparedHP":hpCompared,
                "QuestionCompared":QuestionCompared,
                "percentage":l})
           
        

 


#print (l2)


#print (l3)
#print (l4)
 
#l1, l2 = [x.split(',')[0] for x in originalList], [x.split(',')[1]for x in originalList]
            

#for x in l2:
    #s = SequenceMatcher(None,l2,b).ratio()

#data = {
        #"HP":hp,      
        #"Question":question,  
        #"HPCompared":hpCompared,
        #"QuestionCompared":s}
        


""" Remove the " to allow filter for sentences that match >80
        if (e >= 80):
            #print (c)
            data={"Health Plan":healthPlan,
                "Sentence":a,  
                "Sentnece Compared":b,
                "Percent Match":e}
            df1 = pd.DataFrame.from_dict(data, orient="index")
            df1 = df1.transpose()  
        else :
            data={"Health Plan":healthPlan,
                "Sentence":"",  
                "Sentnece Compared":"",
                "Percent Match":""}  
 """
 
df1 = pd.DataFrame.from_dict(data)       

df1.to_sql('tbl_Sentence_Compare', con=engine, if_exists='replace', index=False)
 

